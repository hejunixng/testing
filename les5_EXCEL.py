# 安装 openpyxl

import openpyxl

'''
三大对象：工作簿workbook、表单sheet、表格cell
1.打开excel文件  load_workbook
2.选中sheet   wb['register']
3.选中单元格 .cell(row,column)
4.获取值 cell1.value
5.写入数据  cell1.value=’xxx‘   ，保存 wb.save('')
'''

'''
wb = openpyxl.load_workbook('test.xlsx')   # 获取excel文件
sheet1 = wb['register']        # 获取文件的register表单
cell1 = sheet1.cell(row=1,column=1)

cell1.value = '用例编号'    # 写入excel文件
wb.save('test.xlsx')        #保存文件

print(cell1.value)
'''

def get_Excel():
    wb = openpyxl.load_workbook('test.xlsx')
    sheet1 = wb['register']
    max_row = sheet1.max_row

    # 放到list里面
    case_list = []
    for i in range(2,max_row):
        case_dict = dict(
            case_id = sheet1.cell(i,1).value,
            url = sheet1.cell(i,5).value,
            data = sheet1.cell(i,6).value,
            expect_data = sheet1.cell(i,7).value
         )
        case_list.append(case_dict)
    print(case_list)
    return case_list
get_Excel()


# 写入Excel
def write_data(filename,sheet,row,column,res_value):
    wb = openpyxl(filename)
    sheet1 = wb[sheet]
    sheet1.cell(row,column).value = res_value
    wb.save(filename)