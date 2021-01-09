"""
接口自动化步骤
1.读取数据
2.发送请求
3.执行结果，对比用例
4.数据写入Excel
"""
import openpyxl
import requests


# 读取
def get_case():
    wb = openpyxl.load_workbook('test.xlsx')
    sheet1 = wb['register']
    max_row = sheet1.max_row

    # 存储到列表
    case_list = []
    for i in range(2, max_row+1):
        case_dict = dict(
            case_id=sheet1.cell(row=i, column=1).value,
            case_url=sheet1.cell(row=i, column=5).value,
            case_data=sheet1.cell(row=i, column=6).value,
            expect_data=sheet1.cell(row=i, column=7).value,
        )
        case_list.append(case_dict)
    return case_list


cases = get_case()
# print(cases)

wb = openpyxl.load_workbook('test.xlsx')
sheet1 = wb['register']

# 发送请求
for case in cases:
    case_id = case['case_id']
    url = case['case_url']
    data = eval(case['case_data'])
    # 用例期望结果
    expect_data = eval(case['expect_data'].replace('null', 'None'))
    expect = expect_data['msg']
    # print(expect)
    # 发送
    response = requests.post(url=url, data=data)
    html = response.json()
    # 获取接口返回的msg 也可以用 code对比
    real_res = html.get('msg')

    if expect == real_res:
        # print('success')
        value='success'
    else:
        # print('fail')
        value = 'fail'
    # 写入Excel
    print(real_res)
    sheet1.cell(row=case_id+1,column=8).value = value
    wb.save('test.xlsx')

